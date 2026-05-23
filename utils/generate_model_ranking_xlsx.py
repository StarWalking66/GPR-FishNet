import argparse
import json
import shutil
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple, Union
from xml.sax.saxutils import escape

import numpy as np


PROJECT_ROOT = Path(__file__).resolve().parents[1]
MODEL_OUTCOMES_DIR = PROJECT_ROOT / "model_outcomes"
DEFAULT_OUTPUT_DIR = MODEL_OUTCOMES_DIR
MASK_PATH = PROJECT_ROOT / "data" / "ST_FishNet_Features" / "all_vars_train_mask_intersection.npy"
THRESHOLD = 0.3175
CHECKPOINT_GLOB = "checkpoints_*"
TIMESTAMP_FMT = "%Y%m%d_%H%M%S"

MODEL_LABELS = {
    "checkpoints_convlstm_baseline": "ConvLSTM",
    "checkpoints_exprecast_final": "ExPreCast",
    "checkpoints_gpr_fishnet_final": "GPR-FishNet (ours)",
    "checkpoints_ksa_predrnn_final": "KSA-PredRNN",
    "checkpoints_pfgnet_final": "PFGNet",
    "checkpoints_predrnn_baseline": "PredRNN",
    "checkpoints_predrnn_v2_baseline": "PredRNN-V2",
    "checkpoints_seacast_baseline": "SeaCast",
    "checkpoints_swinlstm_baseline": "SwinLSTM",
    "checkpoints_timekan_final": "TimeKAN",
}

CONTINUOUS_METRICS: List[Tuple[str, bool]] = [
    ("test_one_step_MAE", False),
    ("test_one_step_MSE", False),
    ("test_one_step_RMSE", False),
    ("test_one_step_R2", True),
    ("test_one_step_SSIM", True),
    ("rollout_2024_MAE", False),
    ("rollout_2024_MSE", False),
    ("rollout_2024_RMSE", False),
    ("rollout_2024_R2", True),
    ("rollout_2024_SSIM", True),
]

def threshold_tag(threshold: float) -> str:
    text = f"{threshold:.10f}".rstrip("0").rstrip(".")
    return text.replace("-", "neg_").replace(".", "p")


def threshold_metric_key(split_prefix: str, metric_name: str, threshold: float) -> str:
    return f"{split_prefix}_{metric_name}_{threshold_tag(threshold)}"


def build_threshold_metric_specs(threshold: float) -> List[Tuple[str, bool]]:
    return [
        (threshold_metric_key("test_one_step", "CSI", threshold), True),
        (threshold_metric_key("test_one_step", "F1", threshold), True),
        (threshold_metric_key("test_one_step", "Precision", threshold), True),
        (threshold_metric_key("test_one_step", "Recall", threshold), True),
        (threshold_metric_key("rollout_2024", "CSI", threshold), True),
        (threshold_metric_key("rollout_2024", "F1", threshold), True),
        (threshold_metric_key("rollout_2024", "Precision", threshold), True),
        (threshold_metric_key("rollout_2024", "Recall", threshold), True),
    ]


THRESHOLD_METRICS: List[Tuple[str, bool]] = build_threshold_metric_specs(THRESHOLD)


def build_artifact_tag(threshold: float, model_count: int, timestamp: str) -> str:
    return f"threshold_{threshold_tag(threshold)}_{model_count}models_{timestamp}"


def infer_model_label(
    checkpoint_dir_name: str,
    label_overrides: Optional[Dict[str, str]] = None,
) -> str:
    merged_labels = dict(MODEL_LABELS)
    if label_overrides:
        merged_labels.update(label_overrides)

    if checkpoint_dir_name in merged_labels:
        return merged_labels[checkpoint_dir_name]

    cleaned = checkpoint_dir_name
    for prefix in ("checkpoints_",):
        if cleaned.startswith(prefix):
            cleaned = cleaned[len(prefix):]
    for suffix in ("_baseline", "_final"):
        if cleaned.endswith(suffix):
            cleaned = cleaned[: -len(suffix)]

    tokens = cleaned.split("_")
    normalized = []
    for token in tokens:
        low = token.lower()
        if low in {"r2", "rmse", "mse", "mae"}:
            normalized.append(token.upper())
        elif low == "v2":
            normalized.append("V2")
        else:
            normalized.append(token.capitalize())
    return " ".join(normalized)


def try_load_openpyxl_deps():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None
    return Workbook, Alignment, Font, PatternFill, get_column_letter


def load_json(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def extract_metrics(payload: Dict[str, Any]) -> Dict[str, Any]:
    metrics = payload.get("metrics")
    if isinstance(metrics, dict):
        return metrics
    return payload


def expand_mask(mask_2d: np.ndarray, shape: Tuple[int, ...]) -> np.ndarray:
    mask = mask_2d.astype(bool)
    while mask.ndim < len(shape):
        mask = np.expand_dims(mask, axis=0)
    return np.broadcast_to(mask, shape)


def compute_threshold_metrics(
    preds: np.ndarray,
    targets: np.ndarray,
    mask_2d: np.ndarray,
    threshold: float,
) -> Dict[str, float]:
    preds = np.asarray(preds, dtype=np.float64)
    targets = np.asarray(targets, dtype=np.float64)

    valid_mask = expand_mask(mask_2d, preds.shape)
    p = preds[valid_mask]
    t = targets[valid_mask]

    if p.size == 0:
        return {"CSI": 0.0, "F1": 0.0, "Precision": 0.0, "Recall": 0.0}

    pred_bin = p >= threshold
    target_bin = t >= threshold

    tp = np.sum(pred_bin & target_bin)
    fp = np.sum(pred_bin & ~target_bin)
    fn = np.sum(~pred_bin & target_bin)

    csi = float(tp / (tp + fp + fn + 1e-8))
    precision = float(tp / (tp + fp + 1e-8))
    recall = float(tp / (tp + fn + 1e-8))
    f1 = float(2.0 * precision * recall / (precision + recall + 1e-8))
    return {"CSI": csi, "F1": f1, "Precision": precision, "Recall": recall}


def aggregate_metric_dicts(rows: List[Dict[str, float]]) -> Tuple[Dict[str, float], Dict[str, float]]:
    if not rows:
        return {}, {}

    keys = list(rows[0].keys())
    means: Dict[str, float] = {}
    stds: Dict[str, float] = {}
    for key in keys:
        values = np.asarray([float(row[key]) for row in rows], dtype=np.float64)
        means[key] = float(np.mean(values))
        stds[key] = float(np.std(values, ddof=0))
    return means, stds


def discover_checkpoint_dirs(
    model_outcomes_dir: Path,
    checkpoint_glob: str = CHECKPOINT_GLOB,
) -> List[Path]:
    return sorted(path for path in model_outcomes_dir.glob(checkpoint_glob) if path.is_dir())


def load_split_metrics(metrics_path: Path) -> Dict[str, float]:
    metric_map = extract_metrics(load_json(metrics_path))
    required_names = ("MAE", "MSE", "RMSE", "R2", "SSIM")
    missing = [name for name in required_names if name not in metric_map]
    if missing:
        raise KeyError(f"{metrics_path.name} missing keys: {', '.join(missing)}")
    return {name: float(metric_map[name]) for name in required_names}


def try_load_threshold_metrics_file(
    metrics_path: Path,
    threshold: float,
) -> Optional[Tuple[Dict[str, float], Dict[str, float]]]:
    if not metrics_path.exists():
        return None

    payload = load_json(metrics_path)
    if "threshold_norm" in payload:
        try:
            payload_threshold = float(payload["threshold_norm"])
        except (TypeError, ValueError):
            return None
        if abs(payload_threshold - threshold) > 1e-8:
            return None

    metric_map = extract_metrics(payload)
    result = {}
    for metric_name in ("CSI", "F1", "Precision", "Recall"):
        if metric_name in metric_map:
            result[metric_name] = float(metric_map[metric_name])
    if len(result) == 4:
        std_payload = payload.get("std", {})
        std_result: Dict[str, float] = {}
        if isinstance(std_payload, dict):
            for metric_name in ("CSI", "F1", "Precision", "Recall"):
                if metric_name in std_payload:
                    try:
                        std_result[metric_name] = float(std_payload[metric_name])
                    except (TypeError, ValueError):
                        continue
        return result, std_result
    return None


def try_compute_threshold_metrics_from_seed_dirs(
    checkpoint_dir: Path,
    split_name: str,
    mask_2d: Optional[np.ndarray],
    threshold: float,
) -> Optional[Tuple[Dict[str, float], Dict[str, float], str]]:
    if mask_2d is None:
        return None

    seed_dirs = sorted(path for path in checkpoint_dir.glob("seed_*") if path.is_dir())
    if not seed_dirs:
        return None

    threshold_rows: List[Dict[str, float]] = []
    used_seed_dirs: List[str] = []
    for seed_dir in seed_dirs:
        preds_path = seed_dir / f"{split_name}_preds.npy"
        targets_path = seed_dir / f"{split_name}_targets.npy"
        if not (preds_path.exists() and targets_path.exists()):
            continue
        threshold_rows.append(
            compute_threshold_metrics(
                np.load(preds_path),
                np.load(targets_path),
                mask_2d,
                threshold,
            )
        )
        used_seed_dirs.append(seed_dir.name)

    if not threshold_rows:
        return None

    metrics_mean, metrics_std = aggregate_metric_dicts(threshold_rows)
    source = f"seed-level recompute from {len(used_seed_dirs)} seed dirs"
    return metrics_mean, metrics_std, source


def resolve_threshold_metrics(
    checkpoint_dir: Path,
    split_name: str,
    mask_2d: Optional[np.ndarray],
    mask_path: Path,
    threshold: float,
) -> Tuple[Optional[Dict[str, float]], Dict[str, float], str]:
    threshold_file = checkpoint_dir / f"{split_name}_metrics_threshold_{threshold_tag(threshold)}.json"
    threshold_payload = try_load_threshold_metrics_file(threshold_file, threshold)
    if threshold_payload is not None:
        threshold_metrics, threshold_stds = threshold_payload
        return threshold_metrics, threshold_stds, threshold_file.name

    seed_level_payload = try_compute_threshold_metrics_from_seed_dirs(
        checkpoint_dir=checkpoint_dir,
        split_name=split_name,
        mask_2d=mask_2d,
        threshold=threshold,
    )
    if seed_level_payload is not None:
        threshold_metrics, threshold_stds, source = seed_level_payload
        return threshold_metrics, threshold_stds, source

    preds_path = checkpoint_dir / f"{split_name}_preds.npy"
    targets_path = checkpoint_dir / f"{split_name}_targets.npy"
    if preds_path.exists() and targets_path.exists() and mask_2d is not None:
        threshold_metrics = compute_threshold_metrics(
            np.load(preds_path),
            np.load(targets_path),
            mask_2d,
            threshold,
        )
        return threshold_metrics, {}, f"{preds_path.name} + {targets_path.name}"

    missing_parts = []
    if threshold_file.exists():
        missing_parts.append(f"{threshold_file.name} missing Precision/Recall or threshold mismatch")
    else:
        missing_parts.append(f"{threshold_file.name} not found")
    if not preds_path.exists():
        missing_parts.append(f"{preds_path.name} not found")
    if not targets_path.exists():
        missing_parts.append(f"{targets_path.name} not found")
    if mask_2d is None:
        missing_parts.append(f"mask file not found: {mask_path}")
    return None, {}, "; ".join(missing_parts)


def load_checkpoint_rows(
    model_outcomes_dir: Path = MODEL_OUTCOMES_DIR,
    mask_path: Path = MASK_PATH,
    threshold: float = THRESHOLD,
    checkpoint_glob: str = CHECKPOINT_GLOB,
    label_overrides: Optional[Dict[str, str]] = None,
    return_skipped: bool = False,
) -> Union[List[Dict[str, Any]], Tuple[List[Dict[str, Any]], List[Dict[str, str]]]]:
    mask_2d = np.load(mask_path).astype(np.float32) if mask_path.exists() else None
    rows: List[Dict[str, Any]] = []
    skipped: List[Dict[str, str]] = []

    test_csi_key = threshold_metric_key("test_one_step", "CSI", threshold)
    test_f1_key = threshold_metric_key("test_one_step", "F1", threshold)
    test_precision_key = threshold_metric_key("test_one_step", "Precision", threshold)
    test_recall_key = threshold_metric_key("test_one_step", "Recall", threshold)
    rollout_csi_key = threshold_metric_key("rollout_2024", "CSI", threshold)
    rollout_f1_key = threshold_metric_key("rollout_2024", "F1", threshold)
    rollout_precision_key = threshold_metric_key("rollout_2024", "Precision", threshold)
    rollout_recall_key = threshold_metric_key("rollout_2024", "Recall", threshold)

    for checkpoint_dir in discover_checkpoint_dirs(model_outcomes_dir, checkpoint_glob):
        test_metrics_path = checkpoint_dir / "test_one_step_metrics.json"
        rollout_metrics_path = checkpoint_dir / "rollout_2024_metrics.json"

        missing_metrics = [
            path.name for path in (test_metrics_path, rollout_metrics_path) if not path.exists()
        ]
        if missing_metrics:
            skipped.append(
                {
                    "model_dir": checkpoint_dir.name,
                    "reason": "missing required metrics files: " + ", ".join(missing_metrics),
                }
            )
            continue

        try:
            test_metrics = load_split_metrics(test_metrics_path)
            rollout_metrics = load_split_metrics(rollout_metrics_path)
        except Exception as exc:  # noqa: BLE001
            skipped.append({"model_dir": checkpoint_dir.name, "reason": f"failed to parse metrics: {exc}"})
            continue

        test_threshold_metrics, test_threshold_stds, test_threshold_source = resolve_threshold_metrics(
            checkpoint_dir,
            "test_one_step",
            mask_2d,
            mask_path,
            threshold,
        )
        rollout_threshold_metrics, rollout_threshold_stds, rollout_threshold_source = resolve_threshold_metrics(
            checkpoint_dir,
            "rollout_2024",
            mask_2d,
            mask_path,
            threshold,
        )

        if test_threshold_metrics is None or rollout_threshold_metrics is None:
            reasons = []
            if test_threshold_metrics is None:
                reasons.append(f"test_one_step threshold metrics unavailable: {test_threshold_source}")
            if rollout_threshold_metrics is None:
                reasons.append(f"rollout_2024 threshold metrics unavailable: {rollout_threshold_source}")
            skipped.append({"model_dir": checkpoint_dir.name, "reason": " | ".join(reasons)})
            continue

        row = {
            "model_label": infer_model_label(checkpoint_dir.name, label_overrides=label_overrides),
            "model_dir": checkpoint_dir.name,
            "threshold_norm": float(threshold),
            "threshold_source_test_one_step": test_threshold_source,
            "threshold_source_rollout_2024": rollout_threshold_source,
            "test_one_step_MAE": test_metrics["MAE"],
            "test_one_step_MSE": test_metrics["MSE"],
            "test_one_step_RMSE": test_metrics["RMSE"],
            "test_one_step_R2": test_metrics["R2"],
            "test_one_step_SSIM": test_metrics["SSIM"],
            "rollout_2024_MAE": rollout_metrics["MAE"],
            "rollout_2024_MSE": rollout_metrics["MSE"],
            "rollout_2024_RMSE": rollout_metrics["RMSE"],
            "rollout_2024_R2": rollout_metrics["R2"],
            "rollout_2024_SSIM": rollout_metrics["SSIM"],
            test_csi_key: test_threshold_metrics["CSI"],
            test_f1_key: test_threshold_metrics["F1"],
            test_precision_key: test_threshold_metrics["Precision"],
            test_recall_key: test_threshold_metrics["Recall"],
            rollout_csi_key: rollout_threshold_metrics["CSI"],
            rollout_f1_key: rollout_threshold_metrics["F1"],
            rollout_precision_key: rollout_threshold_metrics["Precision"],
            rollout_recall_key: rollout_threshold_metrics["Recall"],
        }
        for metric_name, metric_std in test_threshold_stds.items():
            row[f"{threshold_metric_key('test_one_step', metric_name, threshold)}_std"] = float(metric_std)
        for metric_name, metric_std in rollout_threshold_stds.items():
            row[f"{threshold_metric_key('rollout_2024', metric_name, threshold)}_std"] = float(metric_std)
        rows.append(row)

    if return_skipped:
        return rows, skipped
    return rows


def _read_std_payload(path: Path) -> Dict[str, float]:
    if not path.exists():
        return {}
    payload = load_json(path)
    std_payload = payload.get("std")
    if not isinstance(std_payload, dict):
        return {}
    result: Dict[str, float] = {}
    for key, value in std_payload.items():
        try:
            result[key] = float(value)
        except (TypeError, ValueError):
            continue
    return result


def attach_metric_stds(
    rows: List[Dict[str, Any]],
    model_outcomes_dir: Path = MODEL_OUTCOMES_DIR,
    mask_path: Path = MASK_PATH,
    threshold: float = THRESHOLD,
) -> None:
    threshold_file_tag = threshold_tag(threshold)
    mask_2d = np.load(mask_path).astype(np.float32) if mask_path.exists() else None
    for row in rows:
        checkpoint_dir = model_outcomes_dir / row["model_dir"]

        for split_name in ("test_one_step", "rollout_2024"):
            split_std_map = _read_std_payload(checkpoint_dir / f"{split_name}_metrics.json")
            for metric_name, metric_std in split_std_map.items():
                row.setdefault(f"{split_name}_{metric_name}_std", metric_std)

            threshold_std_map = _read_std_payload(checkpoint_dir / f"{split_name}_metrics_threshold_{threshold_file_tag}.json")
            if not threshold_std_map:
                recomputed = try_compute_threshold_metrics_from_seed_dirs(
                    checkpoint_dir=checkpoint_dir,
                    split_name=split_name,
                    mask_2d=mask_2d,
                    threshold=threshold,
                )
                if recomputed is not None:
                    _, threshold_std_map, _ = recomputed
            for metric_name, metric_std in threshold_std_map.items():
                row.setdefault(f"{threshold_metric_key(split_name, metric_name, threshold)}_std", metric_std)


def format_mean_std(value: float, std_value: Optional[float]) -> str:
    if std_value is None:
        return f"{value:.6f}"
    return f"{value:.6f}\u00b1{std_value:.6f}"


def build_display_rows(rows: List[Dict[str, Any]], headers: List[str]) -> List[Dict[str, Any]]:
    display_rows: List[Dict[str, Any]] = []
    for row in rows:
        display_row = dict(row)
        for header in headers:
            std_key = f"{header}_std"
            if std_key not in row:
                continue
            value = row.get(header)
            std_value = row.get(std_key)
            if not isinstance(value, (int, float, np.integer, np.floating)):
                continue
            if not isinstance(std_value, (int, float, np.integer, np.floating)):
                continue
            display_row[header] = format_mean_std(float(value), float(std_value))
        display_rows.append(display_row)
    return display_rows


def assign_ranks(rows: List[Dict[str, Any]], metric_specs: Iterable[Tuple[str, bool]]) -> None:
    for metric_name, higher_is_better in metric_specs:
        sorted_rows = sorted(rows, key=lambda row: row[metric_name], reverse=higher_is_better)
        for index, row in enumerate(sorted_rows, start=1):
            row[f"{metric_name}_rank"] = index


def add_average_ranks(
    rows: List[Dict[str, Any]],
    continuous_metrics: Iterable[Tuple[str, bool]] = CONTINUOUS_METRICS,
    threshold_metrics: Optional[Iterable[Tuple[str, bool]]] = None,
) -> None:
    threshold_metric_specs = list(threshold_metrics or THRESHOLD_METRICS)
    continuous_metric_specs = list(continuous_metrics)

    for row in rows:
        continuous_ranks = [row[f"{name}_rank"] for name, _ in continuous_metric_specs]
        threshold_ranks = [row[f"{name}_rank"] for name, _ in threshold_metric_specs]
        all_ranks = continuous_ranks + threshold_ranks

        row["continuous_avg_rank"] = float(sum(continuous_ranks) / len(continuous_ranks))
        row["threshold_avg_rank"] = float(sum(threshold_ranks) / len(threshold_ranks))
        row["overall_avg_rank"] = float(sum(all_ranks) / len(all_ranks))


def write_sheet(ws: Any, headers: List[str], rows: List[Dict[str, Any]], top_three_col: str = "") -> None:
    deps = try_load_openpyxl_deps()
    if deps is None:
        raise RuntimeError("openpyxl is not available for styled workbook generation.")
    _, Alignment, Font, PatternFill, get_column_letter = deps
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    top1_fill = PatternFill("solid", fgColor="FFF2CC")
    top2_fill = PatternFill("solid", fgColor="EDEDED")
    top3_fill = PatternFill("solid", fgColor="FCE5CD")

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_dict in rows:
        ws.append([row_dict.get(header, "") for header in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    rank_col_index = headers.index(top_three_col) + 1 if top_three_col and top_three_col in headers else None

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, float):
                cell.number_format = "0.000000"
            cell.alignment = Alignment(horizontal="center", vertical="center")

        if rank_col_index is not None:
            rank_value = ws.cell(row=row_idx, column=rank_col_index).value
            if rank_value == 1:
                fill = top1_fill
            elif rank_value == 2:
                fill = top2_fill
            elif rank_value == 3:
                fill = top3_fill
            else:
                fill = None
            if fill is not None:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 24)


def prepare_sheet_specs(
    rows: List[Dict[str, Any]],
    archive_output_path: Path,
    latest_output_path: Path,
    skipped_rows: List[Dict[str, str]],
    threshold: float = THRESHOLD,
    model_outcomes_dir: Path = MODEL_OUTCOMES_DIR,
    mask_path: Path = MASK_PATH,
    checkpoint_glob: str = CHECKPOINT_GLOB,
) -> List[Dict[str, Any]]:
    test_csi_key = threshold_metric_key("test_one_step", "CSI", threshold)
    test_f1_key = threshold_metric_key("test_one_step", "F1", threshold)
    test_precision_key = threshold_metric_key("test_one_step", "Precision", threshold)
    test_recall_key = threshold_metric_key("test_one_step", "Recall", threshold)
    rollout_csi_key = threshold_metric_key("rollout_2024", "CSI", threshold)
    rollout_f1_key = threshold_metric_key("rollout_2024", "F1", threshold)
    rollout_precision_key = threshold_metric_key("rollout_2024", "Precision", threshold)
    rollout_recall_key = threshold_metric_key("rollout_2024", "Recall", threshold)

    overall_rows = sorted(
        rows,
        key=lambda row: (row["overall_avg_rank"], row[f"{rollout_csi_key}_rank"], row[f"{test_csi_key}_rank"]),
    )
    for index, row in enumerate(overall_rows, start=1):
        row["overall_rank"] = index

    continuous_rows = sorted(
        rows,
        key=lambda row: (row["continuous_avg_rank"], row["rollout_2024_R2_rank"], row["test_one_step_R2_rank"]),
    )
    for index, row in enumerate(continuous_rows, start=1):
        row["continuous_rank"] = index

    threshold_rows = sorted(
        rows,
        key=lambda row: (row["threshold_avg_rank"], row[f"{rollout_csi_key}_rank"], row[f"{test_csi_key}_rank"]),
    )
    for index, row in enumerate(threshold_rows, start=1):
        row["threshold_rank"] = index

    one_step_rows = sorted(
        rows,
        key=lambda row: (-row[test_csi_key], -row[test_f1_key], row["test_one_step_MAE"]),
    )
    for index, row in enumerate(one_step_rows, start=1):
        row["one_step_threshold_rank"] = index

    rollout_rows = sorted(
        rows,
        key=lambda row: (-row[rollout_csi_key], -row[rollout_f1_key], row["rollout_2024_MAE"]),
    )
    for index, row in enumerate(rollout_rows, start=1):
        row["rollout_threshold_rank"] = index

    scan_status_rows = [
        {
            "model_dir": row["model_dir"],
            "model_label": row["model_label"],
            "status": "included",
            "reason": "complete metrics",
            "threshold_source_test_one_step": row["threshold_source_test_one_step"],
            "threshold_source_rollout_2024": row["threshold_source_rollout_2024"],
        }
        for row in rows
    ]
    scan_status_rows.extend(
        {
            "model_dir": item["model_dir"],
            "model_label": "",
            "status": "skipped",
            "reason": item["reason"],
            "threshold_source_test_one_step": "",
            "threshold_source_rollout_2024": "",
        }
        for item in skipped_rows
    )
    scan_status_rows.sort(key=lambda item: (item["status"] != "included", item["model_dir"]))

    notes_rows = [
        ("generated_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("archive_output_file", str(archive_output_path)),
        ("latest_output_file", str(latest_output_path)),
        ("source_dir", str(model_outcomes_dir)),
        ("mask_file", str(mask_path)),
        ("checkpoint_glob", checkpoint_glob),
        ("threshold_norm", threshold),
        ("included_model_count", len(rows)),
        ("skipped_checkpoint_count", len(skipped_rows)),
        ("continuous_metrics", "one-step and rollout: MAE, MSE, RMSE, R2, SSIM"),
        ("threshold_metrics", f"recomputed or loaded at threshold {threshold:.4f}: CSI, F1, Precision, Recall"),
        ("overall_rank_rule", "average rank of continuous metrics and threshold metrics on the shared ocean mask"),
        ("display_rule", "metric cells show mean±std when multi-seed summaries provide std; ranks are computed from mean metrics"),
        ("note", "The scan_status sheet lists all discovered checkpoint directories and the reason for inclusion or skipping."),
    ]

    overall_headers = [
        "overall_rank",
        "model_label",
        "model_dir",
        "overall_avg_rank",
        "continuous_avg_rank",
        "threshold_avg_rank",
        "test_one_step_MAE",
        "test_one_step_RMSE",
        "test_one_step_R2",
        "test_one_step_SSIM",
        test_csi_key,
        test_f1_key,
        test_precision_key,
        test_recall_key,
        "rollout_2024_MAE",
        "rollout_2024_RMSE",
        "rollout_2024_R2",
        "rollout_2024_SSIM",
        rollout_csi_key,
        rollout_f1_key,
        rollout_precision_key,
        rollout_recall_key,
    ]
    continuous_headers = [
        "continuous_rank",
        "model_label",
        "model_dir",
        "continuous_avg_rank",
        "test_one_step_MAE",
        "test_one_step_MAE_rank",
        "test_one_step_MSE",
        "test_one_step_MSE_rank",
        "test_one_step_RMSE",
        "test_one_step_RMSE_rank",
        "test_one_step_R2",
        "test_one_step_R2_rank",
        "test_one_step_SSIM",
        "test_one_step_SSIM_rank",
        "rollout_2024_MAE",
        "rollout_2024_MAE_rank",
        "rollout_2024_MSE",
        "rollout_2024_MSE_rank",
        "rollout_2024_RMSE",
        "rollout_2024_RMSE_rank",
        "rollout_2024_R2",
        "rollout_2024_R2_rank",
        "rollout_2024_SSIM",
        "rollout_2024_SSIM_rank",
    ]
    threshold_headers = [
        "threshold_rank",
        "model_label",
        "model_dir",
        "threshold_avg_rank",
        test_csi_key,
        f"{test_csi_key}_rank",
        test_f1_key,
        f"{test_f1_key}_rank",
        test_precision_key,
        f"{test_precision_key}_rank",
        test_recall_key,
        f"{test_recall_key}_rank",
        rollout_csi_key,
        f"{rollout_csi_key}_rank",
        rollout_f1_key,
        f"{rollout_f1_key}_rank",
        rollout_precision_key,
        f"{rollout_precision_key}_rank",
        rollout_recall_key,
        f"{rollout_recall_key}_rank",
    ]
    one_step_headers = [
        "one_step_threshold_rank",
        "model_label",
        "model_dir",
        test_csi_key,
        test_f1_key,
        test_precision_key,
        test_recall_key,
        "test_one_step_MAE",
        "test_one_step_RMSE",
        "test_one_step_R2",
        "test_one_step_SSIM",
    ]
    rollout_headers = [
        "rollout_threshold_rank",
        "model_label",
        "model_dir",
        rollout_csi_key,
        rollout_f1_key,
        rollout_precision_key,
        rollout_recall_key,
        "rollout_2024_MAE",
        "rollout_2024_RMSE",
        "rollout_2024_R2",
        "rollout_2024_SSIM",
    ]

    return [
        {
            "title": "overall_rank",
            "headers": overall_headers,
            "rows": build_display_rows(overall_rows, overall_headers),
            "top_three_col": "overall_rank",
        },
        {
            "title": "continuous_rank",
            "headers": continuous_headers,
            "rows": build_display_rows(continuous_rows, continuous_headers),
            "top_three_col": "continuous_rank",
        },
        {
            "title": "threshold_rank",
            "headers": threshold_headers,
            "rows": build_display_rows(threshold_rows, threshold_headers),
            "top_three_col": "threshold_rank",
        },
        {
            "title": "one_step_threshold",
            "headers": one_step_headers,
            "rows": build_display_rows(one_step_rows, one_step_headers),
            "top_three_col": "one_step_threshold_rank",
        },
        {
            "title": "rollout_threshold",
            "headers": rollout_headers,
            "rows": build_display_rows(rollout_rows, rollout_headers),
            "top_three_col": "rollout_threshold_rank",
        },
        {
            "title": "scan_status",
            "headers": [
                "model_dir",
                "model_label",
                "status",
                "reason",
                "threshold_source_test_one_step",
                "threshold_source_rollout_2024",
            ],
            "rows": scan_status_rows,
            "top_three_col": "",
        },
        {
            "title": "notes",
            "headers": ["key", "value"],
            "rows": [{"key": key, "value": value} for key, value in notes_rows],
            "top_three_col": "",
        },
    ]


def build_workbook(sheet_specs: List[Dict[str, Any]]) -> Any:
    deps = try_load_openpyxl_deps()
    if deps is None:
        raise RuntimeError("openpyxl is not available for styled workbook generation.")
    Workbook, _, _, _, _ = deps
    wb = Workbook()

    first_spec = sheet_specs[0]
    ws = wb.active
    ws.title = first_spec["title"]
    write_sheet(ws, first_spec["headers"], first_spec["rows"], first_spec["top_three_col"])

    for spec in sheet_specs[1:]:
        ws = wb.create_sheet(spec["title"])
        write_sheet(ws, spec["headers"], spec["rows"], spec["top_three_col"])
        if spec["title"] == "notes":
            ws.column_dimensions["A"].width = 24
            ws.column_dimensions["B"].width = 110
        if spec["title"] == "scan_status":
            ws.column_dimensions["A"].width = 34
            ws.column_dimensions["B"].width = 24
            ws.column_dimensions["C"].width = 14
            ws.column_dimensions["D"].width = 80
            ws.column_dimensions["E"].width = 34
            ws.column_dimensions["F"].width = 34

    return wb


def get_column_letter_local(col_idx: int) -> str:
    result = []
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result.append(chr(65 + remainder))
    return "".join(reversed(result))


def value_to_cell_xml(cell_ref: str, value: Any) -> str:
    if value is None:
        return f'<c r="{cell_ref}"/>'
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return f'<c r="{cell_ref}"><v>{value}</v></c>'
    text = escape(str(value))
    return f'<c r="{cell_ref}" t="inlineStr"><is><t>{text}</t></is></c>'


def build_sheet_xml(headers: List[str], rows: List[Dict[str, Any]]) -> str:
    all_rows = [headers] + [[row.get(header, "") for header in headers] for row in rows]
    if not all_rows:
        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData/></worksheet>'
        )

    max_col = get_column_letter_local(len(headers))
    max_row = len(all_rows)
    xml_rows = []
    for row_idx, row_values in enumerate(all_rows, start=1):
        cells = []
        for col_idx, value in enumerate(row_values, start=1):
            cell_ref = f"{get_column_letter_local(col_idx)}{row_idx}"
            cells.append(value_to_cell_xml(cell_ref, value))
        xml_rows.append(f'<row r="{row_idx}">{"".join(cells)}</row>')

    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<dimension ref="A1:{max_col}{max_row}"/>'
        '<sheetViews><sheetView workbookViewId="0"/></sheetViews>'
        '<sheetFormatPr defaultRowHeight="15"/>'
        f'<sheetData>{"".join(xml_rows)}</sheetData>'
        '</worksheet>'
    )


def save_xlsx_fallback(sheet_specs: List[Dict[str, Any]], output_path: Path) -> None:
    timestamp = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheets>'
        + "".join(
            f'<sheet name="{escape(spec["title"])}" sheetId="{idx}" r:id="rId{idx}"/>'
            for idx, spec in enumerate(sheet_specs, start=1)
        )
        + '</sheets></workbook>'
    )

    workbook_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        + "".join(
            f'<Relationship Id="rId{idx}" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
            f'Target="worksheets/sheet{idx}.xml"/>'
            for idx, _ in enumerate(sheet_specs, start=1)
        )
        + '</Relationships>'
    )

    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        + "".join(
            f'<Override PartName="/xl/worksheets/sheet{idx}.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            for idx, _ in enumerate(sheet_specs, start=1)
        )
        + '<Override PartName="/docProps/core.xml" '
        'ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>'
        '<Override PartName="/docProps/app.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>'
        '</Types>'
    )

    root_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>'
        '<Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>'
        '</Relationships>'
    )

    app_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" '
        'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
        '<Application>Python</Application>'
        f'<TitlesOfParts><vt:vector size="{len(sheet_specs)}" baseType="lpstr">'
        + "".join(f'<vt:lpstr>{escape(spec["title"])}</vt:lpstr>' for spec in sheet_specs)
        + '</vt:vector></TitlesOfParts>'
        f'<HeadingPairs><vt:vector size="2" baseType="variant"><vt:variant><vt:lpstr>Worksheets</vt:lpstr></vt:variant><vt:variant><vt:i4>{len(sheet_specs)}</vt:i4></vt:variant></vt:vector></HeadingPairs>'
        '</Properties>'
    )

    core_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
        'xmlns:dc="http://purl.org/dc/elements/1.1/" '
        'xmlns:dcterms="http://purl.org/dc/terms/" '
        'xmlns:dcmitype="http://purl.org/dc/dcmitype/" '
        'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        '<dc:creator>Codex</dc:creator>'
        '<cp:lastModifiedBy>Codex</cp:lastModifiedBy>'
        f'<dcterms:created xsi:type="dcterms:W3CDTF">{timestamp}</dcterms:created>'
        f'<dcterms:modified xsi:type="dcterms:W3CDTF">{timestamp}</dcterms:modified>'
        '</cp:coreProperties>'
    )

    with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", root_rels)
        zf.writestr("docProps/app.xml", app_xml)
        zf.writestr("docProps/core.xml", core_xml)
        zf.writestr("xl/workbook.xml", workbook_xml)
        zf.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
        for idx, spec in enumerate(sheet_specs, start=1):
            zf.writestr(
                f"xl/worksheets/sheet{idx}.xml",
                build_sheet_xml(spec["headers"], spec["rows"]),
            )


def resolve_workbook_paths(
    output_dir: Path,
    explicit_output_path: Optional[Path],
    threshold: float,
    model_count: int,
) -> Tuple[Path, Path]:
    if explicit_output_path is not None:
        return explicit_output_path, explicit_output_path

    timestamp = datetime.now().strftime(TIMESTAMP_FMT)
    archive_path = output_dir / f"model_rankings_{build_artifact_tag(threshold, model_count, timestamp)}.xlsx"
    latest_path = output_dir / f"model_rankings_latest_threshold_{threshold_tag(threshold)}.xlsx"
    return archive_path, latest_path


def save_workbook(sheet_specs: List[Dict[str, Any]], archive_path: Path, latest_path: Path) -> str:
    archive_path.parent.mkdir(parents=True, exist_ok=True)
    latest_path.parent.mkdir(parents=True, exist_ok=True)

    if try_load_openpyxl_deps() is not None:
        wb = build_workbook(sheet_specs)
        wb.save(archive_path)
        writer_mode = "openpyxl"
    else:
        save_xlsx_fallback(sheet_specs, archive_path)
        writer_mode = "stdlib-fallback"

    if latest_path != archive_path:
        shutil.copy2(archive_path, latest_path)
    return writer_mode


def print_scan_report(rows: List[Dict[str, Any]], skipped_rows: List[Dict[str, str]]) -> None:
    print(f"Included checkpoint dirs: {len(rows)}")
    print(f"Skipped checkpoint dirs: {len(skipped_rows)}")
    for item in skipped_rows:
        print(f"  - {item['model_dir']}: {item['reason']}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Scan model_outcomes checkpoint folders and generate a reusable model ranking workbook."
    )
    parser.add_argument(
        "--model-outcomes-dir",
        type=str,
        default=str(MODEL_OUTCOMES_DIR),
        help="Directory containing checkpoint result folders.",
    )
    parser.add_argument(
        "--mask-path",
        type=str,
        default=str(MASK_PATH),
        help="Shared ocean-mask path used for threshold metric recomputation.",
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        default=str(DEFAULT_OUTPUT_DIR),
        help="Directory used when --output-path is not provided.",
    )
    parser.add_argument(
        "--output-path",
        type=str,
        default="",
        help="Optional explicit .xlsx output path. If omitted, the script writes both archive and latest files.",
    )
    parser.add_argument(
        "--threshold",
        type=float,
        default=THRESHOLD,
        help="Threshold used for CSI/F1/Precision/Recall recomputation.",
    )
    parser.add_argument(
        "--checkpoint-glob",
        type=str,
        default=CHECKPOINT_GLOB,
        help="Glob pattern used to discover checkpoint directories.",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Fail if any discovered checkpoint directory is skipped.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    model_outcomes_dir = Path(args.model_outcomes_dir)
    mask_path = Path(args.mask_path)
    output_dir = Path(args.output_dir)
    explicit_output_path = Path(args.output_path) if args.output_path else None

    if not model_outcomes_dir.is_absolute():
        model_outcomes_dir = (PROJECT_ROOT / model_outcomes_dir).resolve()
    if not mask_path.is_absolute():
        mask_path = (PROJECT_ROOT / mask_path).resolve()
    if not output_dir.is_absolute():
        output_dir = (PROJECT_ROOT / output_dir).resolve()
    if explicit_output_path is not None and not explicit_output_path.is_absolute():
        explicit_output_path = (PROJECT_ROOT / explicit_output_path).resolve()

    rows, skipped_rows = load_checkpoint_rows(
        model_outcomes_dir=model_outcomes_dir,
        mask_path=mask_path,
        threshold=args.threshold,
        checkpoint_glob=args.checkpoint_glob,
        return_skipped=True,
    )
    if not rows:
        raise RuntimeError(
            f"No complete checkpoint metrics found under {model_outcomes_dir} with pattern {args.checkpoint_glob}"
        )
    if args.strict and skipped_rows:
        print_scan_report(rows, skipped_rows)
        raise RuntimeError("Strict mode enabled and some checkpoint directories were skipped.")

    attach_metric_stds(rows, model_outcomes_dir=model_outcomes_dir, mask_path=mask_path, threshold=args.threshold)
    threshold_metrics = build_threshold_metric_specs(args.threshold)
    assign_ranks(rows, CONTINUOUS_METRICS)
    assign_ranks(rows, threshold_metrics)
    add_average_ranks(rows, CONTINUOUS_METRICS, threshold_metrics)

    archive_path, latest_path = resolve_workbook_paths(
        output_dir=output_dir,
        explicit_output_path=explicit_output_path,
        threshold=args.threshold,
        model_count=len(rows),
    )
    sheet_specs = prepare_sheet_specs(
        rows,
        archive_output_path=archive_path,
        latest_output_path=latest_path,
        skipped_rows=skipped_rows,
        threshold=args.threshold,
        model_outcomes_dir=model_outcomes_dir,
        mask_path=mask_path,
        checkpoint_glob=args.checkpoint_glob,
    )
    writer_mode = save_workbook(sheet_specs, archive_path, latest_path)

    print(f"Saved workbook archive to: {archive_path}")
    if latest_path != archive_path:
        print(f"Updated latest workbook to: {latest_path}")
    print(f"Writer mode: {writer_mode}")
    print_scan_report(rows, skipped_rows)


if __name__ == "__main__":
    main()
